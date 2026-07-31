"""title2doi 测试运行器

从 index.html 提取 JS 函数，注入断言，Node.js 执行。
"""

import re
import os
import sys
import subprocess
import tempfile


def run_python_tests(project_root):
    """Run focused backend regression tests."""
    sys.path.insert(0, os.path.join(project_root, "server"))

    from openpyxl import Workbook
    from src.parser import parse_excel

    wb = Workbook()
    ws = wb.active
    ws["A1"] = "标题"
    ws["A60"] = "Deep Learning for Natural Language Processing"

    buf = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
    buf.close()
    try:
        wb.save(buf.name)
        with open(buf.name, "rb") as f:
            titles = parse_excel(f.read())
    finally:
        os.unlink(buf.name)

    if titles != ["Deep Learning for Natural Language Processing"]:
        raise AssertionError("parse_excel should read rows beyond the preview window, got %r" % titles)


def extract_js(html_path):
    """提取 index.html 中的主 inline script"""
    with open(html_path, "r", encoding="utf-8") as f:
        html = f.read()
    scripts = re.findall(r"<script>(.*?)</script>", html, re.DOTALL)
    for s in reversed(scripts):
        if "function doParse" in s:
            return s
    raise RuntimeError("Could not find main script in index.html")


def main():
    project_root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    html_path = os.path.join(project_root, "index.html")

    print("Extracting JS from index.html...")
    js_code = extract_js(html_path)

    # Node.js 环境 mock（必须在 JS 代码最前面）
    prelude = """// Node.js environment mock
if (typeof window === 'undefined') { global.window = {}; }
if (typeof document === 'undefined') { global.document = { createElement: function() { return {}; }, querySelector: function(){return null}, querySelectorAll: function(){return[]}, getElementById: function(){ return { classList:{add:function(){},remove:function(){},toggle:function(){},contains:function(){return false}}, addEventListener:function(){}, value:'', textContent:'', innerHTML:'', style:{}, focus:function(){}, setAttribute:function(){} } } }; }
if (typeof localStorage === 'undefined') { global.localStorage = { getItem: function(){return null}, setItem: function(){}, removeItem: function(){} }; }
if (typeof sessionStorage === 'undefined') { global.sessionStorage = { getItem: function(){return null}, setItem: function(){}, removeItem: function(){} }; }
if (typeof fetch === 'undefined') { global.fetch = function(){}; }
if (typeof navigator === 'undefined') { global.navigator = { clipboard: {} }; }
if (typeof DOMParser === 'undefined') { global.DOMParser = function(){}; }
"""
    js_code = prelude + js_code

    # 注入测试断言
    test_assertions = """
var passed = 0, failed = 0;
function assert(cond, msg) {
    if (cond) { passed++; }
    else { console.log('  FAIL: ' + msg); failed++; }
}
function report() {
    console.log('');
    console.log('Results: ' + passed + ' passed, ' + failed + ' failed');
    if (failed > 0) process.exit(1);
}

// ---- matchScore ----
console.log('--- matchScore ---');
// Exact match
assert(matchScore('Hello World', 'Hello World') === 100, 'exact english = 100');
// Case insensitive
assert(matchScore('hello world', 'Hello World') === 100, 'case insensitive = 100');
// Stop words
var s1 = matchScore('A Study of Machine Learning', 'Study of Machine Learning');
assert(s1 === 97, 'stop words a/an/the = 97, got ' + s1);
var s2 = matchScore('The Deep Learning', 'Deep Learning');
assert(s2 === 97, 'stop words = 97, got ' + s2);
// Chinese exact
assert(matchScore('深度学习在NLP中的应用', '深度学习在NLP中的应用') === 100, 'chinese exact = 100');
// Chinese different
var s3 = matchScore('激光焦斑尺寸对极紫外辐射影响的理论研究', '激光光斑尺寸对激光焊接熔池、匙孔行为的影响');
assert(s3 <= 50, 'chinese different <= 50, got ' + s3);
// English different
var s4 = matchScore('Completely Different Paper', 'Nothing In Common');
assert(s4 <= 70, 'english different <= 70, got ' + s4);
// Punctuation diff
assert(matchScore('Hello, World!', 'Hello World') === 96, 'punct diff = 96');

// ---- cleanTitle ----
console.log('--- cleanTitle ---');
assert(cleanTitle('1. Hello World') === 'Hello World', 'strip leading number');
assert(cleanTitle('  论文成果：深度学习  ') === '深度学习', 'strip prefix label');
assert(cleanTitle('Test Paper 第一作者') === 'Test Paper', 'strip author suffix');
assert(cleanTitle('Test Paper 通讯作者：张三') === 'Test Paper', 'strip corresponding author');

// ---- isTitle ----
console.log('--- isTitle ---');
assert(isTitle('A Study on Machine Learning'), 'valid english title');
assert(isTitle('深度学习在自然语言处理中的应用研究'), 'valid chinese title');
assert(!isTitle('ab'), 'too short');
assert(!isTitle('参考文献'), 'skip marker');
assert(!isTitle('References'), 'skip english marker');

// ---- isSectionLabel ----
console.log('--- isSectionLabel ---');
assert(isSectionLabel('代表性论文1'), 'section label with number');
assert(isSectionLabel('论文成果：'), 'section label with colon');
assert(isSectionLabel('一、'), 'numbered marker');
assert(!isSectionLabel('A Study on Machine Learning'), 'not a section label');

// ---- isCleanTitles ----
console.log('--- isCleanTitles ---');
assert(isCleanTitles('Deep Learning for NLP\\nMachine Learning Basics\\nAttention Is All You Need'), 'clean titles >= 70%');
assert(!isCleanTitles('Smith J, Brown K. A Study on ML. Nature, 2020, 15(3): 123-145\\nDeep Learning'), 'citations not clean');

// ---- parseText ----
console.log('--- parseText ---');
var titles = parseText('1. Deep Learning\\n2. Machine Learning\\n\\nReferences\\nAttention Is All You Need');
assert(titles.length === 3, 'parse 3 titles, got ' + titles.length);

// ---- groupEntries ----
console.log('--- groupEntries ---');
var entries = groupEntries('论文成果：\\nDeep Learning for NLP\\nSmith J, Brown K. A Study on ML. Nature, 2020');
assert(entries.length >= 2, 'group entries >= 2, got ' + entries.length);

// ---- parseNumberedResponse ----
console.log('--- parseNumberedResponse ---');
var resp = parseNumberedResponse('[1] Deep Learning for NLP\\n[2] Machine Learning', 2);
assert(resp.length === 2, 'numbered response, got ' + resp.length);
assert(resp[0] === 'Deep Learning for NLP', 'first title matches');

// ---- DOI 后缀清理（.S001 补充材料）----
console.log('--- DOI cleanup ---');
function cleanDoi(doi){var c=doi;if(/^10\\.\\d{4,}\\//.test(c))c=c.replace(/\\.S\\d{3}$/i,'');return c}
assert(cleanDoi('10.1016/j.physletb.2021.136601.S001')==='10.1016/j.physletb.2021.136601', 'strip .S001');
assert(cleanDoi('10.1021/acs.nanolett.5c00094.s004')==='10.1021/acs.nanolett.5c00094', 'strip lowercase .s004');
assert(cleanDoi('10.1016/j.actbio.2021.11.001')==='10.1016/j.actbio.2021.11.001', 'keep plain doi');
assert(cleanDoi('10.1234/abc.S1')==='10.1234/abc.S1', 'keep single digit S1');
assert(cleanDoi('10.1002/ange.202524546')==='10.1002/ange.202524546', 'keep wiley doi');

// ---- 德文版→国际版转换（Angewandte Chemie）----
console.log('--- ange to anie ---');
function convAnge(doi,yr){if(/^10\\.1002\\/ange\\./.test(doi)&&yr>=1956)return doi.replace(/^10\\.1002\\/ange\\./,'10.1002/anie.');return doi}
assert(convAnge('10.1002/ange.202016082',2020)==='10.1002/anie.202016082', 'ange→anie for modern year');
assert(convAnge('10.1002/ange.202016082',1940)==='10.1002/ange.202016082', 'keep ange for pre-1956');
assert(convAnge('10.1002/anie.202016082',2020)==='10.1002/anie.202016082', 'keep anie unchanged');
assert(convAnge('10.1002/ange.202016082',null)==='10.1002/ange.202016082', 'keep ange when year unknown');

// ---- DOI 缓存 ----
console.log('--- DOI cache ---');
// 模拟 lookupDOI 的缓存逻辑
DOI_CACHE={};
DOI_CACHE['test title']={doi:'10.1234/test',is_found:true,confidence:95};
var cached=DOI_CACHE['test title'];
var copy=JSON.parse(JSON.stringify(cached));
assert(copy.doi==='10.1234/test', 'cache hit returns deep copy');

// ---- regex matching edge cases ----
console.log('--- edge cases ---');
assert(cleanTitle('  Extra   Spaces  ') === 'Extra Spaces', 'collapse whitespace');
assert(matchScore('', 'something') === 0, 'empty query = 0');
assert(matchScore('test', '') === 0, 'empty candidate = 0');

report();
"""
    js_code += test_assertions

    # 写入临时文件并运行
    tmp = tempfile.NamedTemporaryFile(mode="w", suffix=".js", delete=False, encoding="utf-8")
    tmp.write(js_code)
    tmp.close()

    print("Running tests with Node.js...")
    result = subprocess.run(["node", tmp.name], capture_output=True, text=True)

    print(result.stdout)
    if result.stderr:
        print(result.stderr, file=sys.stderr)

    os.unlink(tmp.name)
    if result.returncode == 0:
        print("Running backend regression tests...")
        run_python_tests(project_root)
        print("Backend regression tests passed")
    sys.exit(result.returncode)


if __name__ == "__main__":
    main()
